import HR_GPT_reply as HR_reply
from openpyxl import load_workbook
from openpyxl import Workbook

from typing import Final
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes



TOKEN: Final = 'TYPE YOUR TELEGRAM API'
BOT_USERNAME: Final = '@TYPE THE NAME OF YOUR TELEGRAM BOT'

def handle_response(text: str, ID: int) -> str:
    # The function provides response from ChatGPT
    #   text - the sent text by a telegram user
    #   ID - Name of a telegram user


    response = HR_reply.HR_GPT_reply_fcn(text, ID)

    return response

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # The function sends messages via telegram

    # The message type classification: private chat or group chat
    message_type: str = update.message.chat.type

    # The sent text by a telegram user
    text: str = update.message.text

    # Print: Username, chat type, user's request
    print(f'User {update.message.chat.username} in {message_type}: "{text}"')

    # Retrieve user's prompt with different techniques based on the chat type (private char or group chat)
    if message_type == 'group':
        # Consider prompts, when users mention the HR bot's username in group chats
        if BOT_USERNAME in text:
            new_text: str = text.replace(BOT_USERNAME, '').strip()
            response: str = handle_response(new_text, update.message.chat.username)
        else:
            return
    else:
        # Retrieve user's prompt, when the users send prompts directly to the HR bot
        response: str = handle_response(text, update.message.chat.username)

    # Provides response from ChatGPT
    await update.message.reply_text(response)

async def error(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # The function prints error on a python console

    print(f'Update {update} caused error {context.error}')

# Print the launch of the bot
print('Starting bot...')

# Connect to your telegram bot
app = Application.builder().token(TOKEN).build()

# Add the response function to your bot
app.add_handler(MessageHandler(filters.TEXT, handle_message))

# Add the error function to your bot
app.add_error_handler(error)

# Set wait time for the bot's response to 1 seconds
app.run_polling(poll_interval =1)





'''

###

wb = load_workbook(filename = 'history_book.xlsx')
#wb = Workbook()
ws = wb.active
#ws.cell(row=4, column=2).value = 'Hi'

match_ID = 2

# Count rows
row_len = len([row for row in ws if not all([cell.value == None for cell in row])])
print(row_len)

# Find row
for row in range(1,row_len):
    if ws.cell(row=row, column=1).value == match_ID:
        print(ws.cell(row=row, column=2).value)

wb.save('history_book.xlsx')
'''