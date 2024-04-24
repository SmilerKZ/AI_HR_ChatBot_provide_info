import GPT_cmd as GC
from openpyxl import load_workbook

def HR_GPT_reply_fcn(prompt, ID):
    # The function receives response from ChatGPT by considering the user's past requests
    # and save the response in xlsx database
    #   prompt - user's prompt
    #   ID - user's telegram username


    filename_history = 'history_book.xlsx'

    filename_system = 'system_prompt.txt'

    context = ''
    system_prompt = ''

    # Read system prompt for the HR bot
    with open(filename_system, encoding="utf8") as file_object:
        for line in file_object:
            system_prompt += line


    # Load xlsx database of users' requests
    wb = load_workbook(filename=filename_history)
    ws = wb.active

    match_ID = ID
    found_flag = False
    found_row = None

    # Count users in the xlsx database of users' requests
    row_len = len([row for row in ws if not all([cell.value == None for cell in row])])

    # Find if the user has sent requests before
    for row in range(1, row_len+1):
        if match_ID == ws.cell(row=row, column=1).value:
            context = ws.cell(row=row, column=2).value
            found_flag = True
            found_row = row

            print("Found")

    # Receive response from ChatGPT on users' requests
    response = GC.GPT_cmd_fcn(prompt, context, system_prompt)

    # Print a user's request on a console
    print(response)

    # Save the user's request and the ChatGPT's response in xlsx database
    new_context = context + 'User: ' + prompt + '\n'
    new_context = new_context + 'ChatGPT: ' + response + '\n'
    if found_flag == True:
        ws.cell(row=found_row, column=2).value = new_context
    else:
        ws.cell(row=row_len+1, column=1).value = ID
        ws.cell(row=row_len+1, column=2).value = new_context

    wb.save('history_book.xlsx')

    return response
